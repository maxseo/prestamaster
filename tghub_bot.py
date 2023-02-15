import asyncio
from telethon import TelegramClient
from telethon.errors import SessionPasswordNeededError
from telethon.errors import rpcerrorlist
from telethon.tl.functions.messages import GetHistoryRequest
from telethon.tl.functions.channels import GetFullChannelRequest
from telethon.tl.types import ChannelParticipantsRecent
import openpyxl


async def main():
    # Enter your own API ID and API hash here
    api_id = 12345
    api_hash = "your_api_hash"
    session_name = 'tghub_bot'
    client = TelegramClient(session_name, api_id, api_hash)
    try:
        await client.start()
    except SessionPasswordNeededError:
        pw = input("Two-factor authentication is enabled. Please enter your password: ")
        await client.start(password=pw)
    except rpcerrorlist.ApiIdInvalidError:
        print("Invalid API ID. Please enter a valid API ID.")
        return
    except rpcerrorlist.AuthKeyError:
        print("Invalid API hash. Please enter a valid API hash.")
        return

    # Open channels.txt file and read channel ids
    with open('channels.txt') as f:
        channels = f.readlines()
    channels = [channel.strip() for channel in channels]

    # Create a new excel file and set header cells
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1, value="Channel ID")
    sheet.cell(row=1, column=2, value="Last 5 Posts")

    # Loop through each channel and get last 5 posts
    for row, channel_id in enumerate(channels, start=2):
        try:
            channel = await client.get_entity(channel_id)
            full_channel = await client(GetFullChannelRequest(channel=channel))
            participants = await client.get_participants(channel, limit=0, offset=0, search="")
            messages = await client(GetHistoryRequest(peer=channel_id, limit=5, offset_date=None, offset_id=0, max_id=0, min_id=0, add_offset=0, hash=0))
            posts = []
            for message in reversed(messages.messages):
                # Check if the message is a media type (i.e. photo, video, document, etc.)
                if message.media is not None:
                    # If media type is photo
                    if message.photo is not None:
                        media = message.photo
                        caption = message.message
                    # If media type is video
                    elif message.video is not None:
                        media = message.video
                        caption = message.message
                    # If media type is document
                    elif message.document is not None:
                        media = message.document
                        caption = message.message
                    # If media type is not photo, video, or document, then we don't include it in the post
                    else:
                        continue
                    post = f'<div class="tg-post"><a href="https://t.me/{channel_id}/{message.id}">{media}</a><div class="tg-text">{caption}</div></div>'
                    posts.append(post)
                # If message is not a media type, then we don't include it in the post
                else:
                    continue
            # Add channel id and its last 5 posts to excel file
            posts_str = "<br>".join(posts)
        wb.save('posts.xlsx')
        print("Excel file created successfully!")

        # Close the client connection
        await client.disconnect()

    if __name__ == '__main__':
        asyncio.run(main())
